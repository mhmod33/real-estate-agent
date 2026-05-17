import os
import json
import time

from dotenv import load_dotenv
from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

client = Groq(api_key=os.getenv("GROQ_API_KEY"))

MODEL_NAME = "llama-3.3-70b-versatile"

AREAS = [
    # القاهرة الكبرى
    "التجمع الخامس", "المعادي", "الزمالك", "مدينة نصر", "المهندسين",
    "الدقي", "العجوزة", "المقطم", "حلوان", "شبرا",
    "المطرية", "عين شمس", "النزهة", "مصر الجديدة", "بولاق",
    "التبين", "دار السلام", "البساتين", "السيدة زينب", "وسط البلد",
    # الجيزة
    "الشيخ زايد", "أكتوبر", "الهرم", "فيصل", "إمبابة",
    "البدرشين", "الحوامدية",
    # المدن الجديدة
    "العاصمة الإدارية الجديدة", "الرحاب", "مدينتي", "الشروق", "بدر",
    "المستقبل سيتي", "العبور", "العلمين الجديدة", "هليوبوليس الجديدة",
    # الإسكندرية
    "سموحة", "المنتزه", "العجمي", "محرم بك", "الإسكندرية وسط",
    "ستانلي", "سيدي بشر", "الإسكندرية الجديدة",
    # الدلتا والقناة
    "المنصورة", "طنطا", "الزقازيق", "بورسعيد", "الإسماعيلية",
    "السويس", "دمياط", "كفر الشيخ",
    # الصعيد
    "أسيوط", "سوهاج", "الأقصر", "أسوان", "المنيا",
    "بني سويف", "الفيوم", "قنا",
]

BATCH_PROMPT_TEMPLATE = """
أنت خبير عقاري في السوق المصري. أنا محتاج بيانات تقريبية عن المناطق دي:
{areas_list}

رجّعلي JSON Array فقط (بدون أي كلام تاني) بالشكل ده:
[
  {{
    "اسم_المنطقة": "",
    "المحافظة": "",
    "المدينة": "",
    "متوسط_سعر_المتر_بيع": 0,
    "متوسط_الإيجار_الشهري": 0,
    "نوع_العقارات_الغالبة": "",
    "تصنيف_المنطقة": "",
    "ملاحظات": ""
  }}
]

ملاحظات:
- متوسط_سعر_المتر_بيع: بالجنيه المصري (رقم فقط)
- متوسط_الإيجار_الشهري: لشقة متوسطة 100-150 متر بالجنيه المصري (رقم فقط)
- نوع_العقارات_الغالبة: مثل "شقق"، "فيلل"، "شقق وفيلل"
- المحافظة: مثل "القاهرة"، "الجيزة"، "الإسكندرية"، "أسيوط"
- المدينة: مثل "القاهرة الجديدة"، "القاهرة"، "الجيزة"
- تصنيف_المنطقة: "راقية" أو "فوق المتوسط" أو "متوسطة" أو "شعبية"
- ملاحظات: جملة قصيرة عن المنطقة (موقعها، مميزاتها، نوع السكان)

رجّع JSON Array فقط بدون أي نص إضافي. لازم يكون عنصر واحد لكل منطقة.
"""

PROGRESS_FILE = os.path.join("data", "_progress.json")


def clean_json(raw: str) -> str:
    """تنظيف الرد من markdown وأي نص زيادة."""
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]
    raw = raw.strip()
    return raw


def load_progress() -> list[dict]:
    """تحميل التقدم المحفوظ."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_progress(data: list[dict]):
    """حفظ التقدم."""
    os.makedirs(os.path.dirname(PROGRESS_FILE), exist_ok=True)
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def fetch_batch(areas: list[str], batch_num: int, total_batches: int) -> list[dict]:
    """جلب بيانات مجموعة مناطق في طلب واحد."""
    areas_list = "\n".join(f"- {a}" for a in areas)
    print(f"\n[Batch {batch_num}/{total_batches}] 🔍 جاري البحث عن {len(areas)} منطقة...")

    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {
                    "role": "system",
                    "content": "أنت خبير عقاري مصري. رجّع JSON Array فقط بدون أي كلام.",
                },
                {
                    "role": "user",
                    "content": BATCH_PROMPT_TEMPLATE.format(areas_list=areas_list),
                },
            ],
            temperature=0.3,
            max_tokens=4096,
        )

        raw = clean_json(response.choices[0].message.content)
        results = json.loads(raw)

        if isinstance(results, list):
            for r in results:
                name = r.get("اسم_المنطقة", "?")
                price = r.get("متوسط_سعر_المتر_بيع", 0)
                print(f"  ✅ {name} — سعر المتر: {price:,} جنيه")
            return results
        else:
            print("  ⚠️  الرد مش JSON Array")
            return []

    except json.JSONDecodeError as e:
        print(f"  ⚠️  فشل تحويل JSON: {e}")
        return []
    except Exception as e:
        print(f"  ❌ خطأ: {e}")
        return []


def create_excel(data: list[dict], output_path: str):
    """إنشاء ملف Excel من البيانات."""
    wb = Workbook()
    ws = wb.active
    ws.title = "أسعار المناطق"
    ws.sheet_view.rightToLeft = True

    # الأعمدة
    headers = [
        "اسم المنطقة",
        "المحافظة",
        "المدينة",
        "متوسط سعر المتر (بيع)",
        "متوسط الإيجار الشهري",
        "نوع العقارات الغالبة",
        "تصنيف المنطقة",
        "ملاحظات",
    ]

    # تنسيق الهيدر
    header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # كتابة الهيدر
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # تنسيق البيانات
    data_font = Font(name="Arial", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(data, 2):
        values = [
            item.get("اسم_المنطقة", ""),
            item.get("المحافظة", ""),
            item.get("المدينة", ""),
            item.get("متوسط_سعر_المتر_بيع", 0),
            item.get("متوسط_الإيجار_الشهري", 0),
            item.get("نوع_العقارات_الغالبة", ""),
            item.get("تصنيف_المنطقة", ""),
            item.get("ملاحظات", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            # تنسيق الأرقام بفاصلة
            if col in (4, 5) and isinstance(value, (int, float)):
                cell.number_format = "#,##0"

    # عرض الأعمدة
    col_widths = [20, 18, 20, 25, 25, 22, 18, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # حفظ
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def main():
    print("=" * 60)
    print("  📊 جمع بيانات أسعار المناطق العقارية (50+ منطقة)")
    print("=" * 60)

    # تحميل التقدم السابق
    all_data = load_progress()
    done_areas = {d.get("اسم_المنطقة") for d in all_data}
    remaining = [a for a in AREAS if a not in done_areas]

    if all_data:
        print(f"\n  📂 تم تحميل {len(all_data)} منطقة من التقدم السابق")

    if not remaining:
        print("  ✅ كل المناطق اتجمعت بالفعل!")
    else:
        print(f"  📋 متبقي {len(remaining)} منطقة")

        # تقسيم لمجموعات (5 مناطق لكل طلب)
        batch_size = 5
        batches = [remaining[i:i + batch_size] for i in range(0, len(remaining), batch_size)]

        for idx, batch in enumerate(batches, 1):
            results = fetch_batch(batch, idx, len(batches))
            all_data.extend(results)

            # حفظ التقدم بعد كل batch
            save_progress(all_data)
            print(f"  💾 تم حفظ التقدم ({len(all_data)}/{len(AREAS)})")

            # تأخير بين الطلبات
            if idx < len(batches):
                time.sleep(2)

    if all_data:
        output_path = os.path.join("data", "areas_prices.xlsx")
        create_excel(all_data, output_path)

        # تنظيف ملف التقدم
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)

        print(f"\n{'=' * 60}")
        print(f"  ✅ تم حفظ {len(all_data)} منطقة في: {output_path}")
        print(f"{'=' * 60}")
    else:
        print("\n❌ مفيش بيانات اتجمعت!")


if __name__ == "__main__":
    main()
