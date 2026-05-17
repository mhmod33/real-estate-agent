"""
تحويل بيانات المناطق العقارية لـ embeddings باستخدام sentence-transformers.
يقرأ من JSON و Excel ويحول كل منطقة لنص واضح ثم يعمل embedding.
"""

import os
import json

from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer

MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
MODEL_PATH = os.path.join(BASE_DIR, "models", "multilingual-minilm")
JSON_FILE = os.path.join(DATA_DIR, "_progress.json")
EXCEL_FILE = os.path.join(DATA_DIR, "areas_prices.xlsx")

_model = None


def get_model() -> SentenceTransformer:
    """تحميل الموديل (مرة واحدة بس)."""
    global _model
    if _model is None:
        if os.path.exists(MODEL_PATH):
            print(f"📦 تحميل الموديل من المسار المحلي: {MODEL_PATH}")
            _model = SentenceTransformer(MODEL_PATH)
        else:
            print(f"📦 جاري تحميل موديل: {MODEL_NAME} ...")
            _model = SentenceTransformer(MODEL_NAME)
            os.makedirs(MODEL_PATH, exist_ok=True)
            _model.save(MODEL_PATH)
            print(f"✅ تم حفظ الموديل في: {MODEL_PATH}")
        print("✅ تم تحميل الموديل")
    return _model




def init_model():
    """Initialize the model at startup."""
    get_model()

def load_from_json(path: str = JSON_FILE) -> list[dict]:
    """قراءة البيانات من ملف JSON."""
    if not os.path.exists(path):
        print(f"⚠️  ملف JSON مش موجود: {path}")
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"📄 تم تحميل {len(data)} منطقة من JSON")
    return data


def load_from_excel(path: str = EXCEL_FILE) -> list[dict]:
    """قراءة البيانات من ملف Excel."""
    if not os.path.exists(path):
        print(f"⚠️  ملف Excel مش موجود: {path}")
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return []

    headers = rows[0]
    data = []
    for row in rows[1:]:
        item = {}
        for h, v in zip(headers, row):
            if h and v is not None:
                key = str(h).strip().replace(" ", "_").replace("(", "").replace(")", "")
                item[key] = v
        if item:
            data.append(item)

    wb.close()
    print(f"📊 تم تحميل {len(data)} منطقة من Excel")
    return data


def normalize_record(record: dict) -> dict:
    """توحيد أسماء الحقول بين JSON و Excel."""
    mapping = {
        "اسم_المنطقة": ["اسم_المنطقة", "اسم_المنطقة"],
        "المحافظة": ["المحافظة"],
        "المدينة": ["المدينة"],
        "متوسط_سعر_المتر_بيع": ["متوسط_سعر_المتر_بيع", "متوسط_سعر_المتر_بيع"],
        "متوسط_الإيجار_الشهري": ["متوسط_الإيجار_الشهري", "متوسط_الإيجار_الشهري"],
        "نوع_العقارات_الغالبة": ["نوع_العقارات_الغالبة", "نوع_العقارات_الغالبة"],
        "تصنيف_المنطقة": ["تصنيف_المنطقة", "تصنيف_المنطقة"],
        "ملاحظات": ["ملاحظات"],
    }

    normalized = {}
    for target_key, source_keys in mapping.items():
        for sk in source_keys:
            if sk in record:
                normalized[target_key] = record[sk]
                break
        if target_key not in normalized:
            normalized[target_key] = record.get(target_key, "")

    return normalized


def record_to_text(record: dict) -> str:
    """تحويل سجل منطقة لنص واضح بالعربي مناسب للـ embedding."""
    r = normalize_record(record)

    name = r.get("اسم_المنطقة", "غير معروف")
    gov = r.get("المحافظة", "")
    city = r.get("المدينة", "")
    price_sqm = r.get("متوسط_سعر_المتر_بيع", 0)
    rent = r.get("متوسط_الإيجار_الشهري", 0)
    prop_type = r.get("نوع_العقارات_الغالبة", "")
    category = r.get("تصنيف_المنطقة", "")
    notes = r.get("ملاحظات", "")

    parts = [f"منطقة {name}"]
    if gov:
        parts.append(f"في محافظة {gov}")
    if city and city != gov:
        parts.append(f"مدينة {city}")
    if price_sqm:
        parts.append(f"متوسط سعر المتر {price_sqm:,} جنيه مصري")
    if rent:
        parts.append(f"متوسط الإيجار الشهري {rent:,} جنيه")
    if prop_type:
        parts.append(f"نوع العقارات الغالبة {prop_type}")
    if category:
        parts.append(f"تصنيف المنطقة {category}")
    if notes:
        parts.append(notes)

    return ". ".join(parts) + "."


def load_all_records() -> list[dict]:
    """تحميل كل البيانات من JSON و Excel بدون تكرار."""
    json_data = load_from_json()
    excel_data = load_from_excel()

    # دمج بدون تكرار (بناءً على اسم المنطقة)
    seen = set()
    all_records = []

    for record in json_data + excel_data:
        r = normalize_record(record)
        name = r.get("اسم_المنطقة", "")
        if name and name not in seen:
            seen.add(name)
            all_records.append(r)

    print(f"📋 إجمالي المناطق بعد الدمج: {len(all_records)}")
    return all_records


def create_embeddings(records: list[dict] | None = None) -> tuple[list[dict], list[str], list[list[float]]]:
    """
    إنشاء embeddings لكل المناطق.

    Returns:
        (records, texts, embeddings)
    """
    if records is None:
        records = load_all_records()

    texts = [record_to_text(r) for r in records]

    model = get_model()
    print(f"🔄 جاري عمل embeddings لـ {len(texts)} منطقة...")
    embeddings = model.encode(texts, show_progress_bar=True, convert_to_numpy=True)

    print(f"✅ تم إنشاء {len(embeddings)} embedding (بُعد: {embeddings.shape[1]})")
    return records, texts, embeddings.tolist()


if __name__ == "__main__":
    records, texts, embeddings = create_embeddings()
    print(f"\n--- عينة ---")
    for i in range(min(3, len(texts))):
        print(f"\n📍 {records[i].get('اسم_المنطقة')}:")
        print(f"   النص: {texts[i][:100]}...")
        print(f"   Embedding dim: {len(embeddings[i])}")
